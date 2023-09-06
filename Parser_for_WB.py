# -*- coding: utf-8 -*
import re
import time
from urllib.parse import urlencode
import logging
import requests

from web_driver_for_selenium import lib
import random

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from settings import TOKEN_ICON


class WildBerriesScraper:
    def __init__(self):
        self.users = lib.USER_AGENTS_PROXY_LIST
        self.number = 1
        self.list_brand = []

    def find_brends(self):
        token = TOKEN_ICON
        # url_brands = 'https://iconmarket.ru/products'
        url_brands = 'https://iconmarket.ru/marketplace/list?filters%5Bsearch%5D=&filters%5Bsub_brand%5D=&filters%5Bcategory_id%5D%5B%5D=15&filters%5Bstop_marketplace%5D=0'
        headers = {'Authorization': f'Bearer {token}'}
        # response = requests.get(url_brands, headers=headers)
        # self.list_brand = re.findall(r'<p style="font-size: 12px">([\w\d\s]+)</p>', response.text)
        # return self.list_brand
        driver_find = self._init_driver()
        driver_find.get(url_brands)
        wait = WebDriverWait(driver_find, 10)
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "brands")))
        driver_find.execute_script("return Object.assign(navigator, {webdriver: false});")
        html = driver_find.get(url_brands, headers=headers)
        soup = BeautifulSoup(html, 'html.parser')
        soup.find("ul", class_="filters_list")
        print(soup.text)

    def run(self):
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s [%(levelname)s]: %(message)s',
            handlers=[logging.FileHandler('scraper.log'), logging.StreamHandler()]
        )

        wb = Workbook()
        ws = wb.active
        self.settings_fro_execel_file(ws)
        ws.append(["номер", "бренд", "название", "цена", "количество отзывов", "ссылка на карточку"])
        self.find_brends()
        for name in range(len(self.list_brand)):
            self.list_brand[name] = self.list_brand[name].replace(' ', '+')
            encoded_brand = urlencode({'search': self.list_brand[name]})
            link = f'https://www.wildberries.ru/catalog/0/search.aspx?page=1&{encoded_brand}'
            driver = self._init_driver()
            try:
                self.all_scrap_func(driver, link, ws)
            except:
                continue
            logging.info(f"сбор данных со страницы -  {link}")
            page = 1
            prev_url = None
            while True:
                driver = self._init_driver()
                try:
                    self.all_scrap_func(driver, link, ws)
                except:
                    break
                page += 1
                logging.info(f"сбор данных со страницы {link}")
                next_link = link.replace(f'page={page - 1}', f'page={page}')
                if next_link == link:  # no more pages
                    page = 1
                    break
                # Check if we are loading the same page again
                prev_url = driver.current_url
                driver.get(next_link)
                if driver.current_url == prev_url:
                    print("Loaded the same page again!")
                    break
                link = next_link

        wb.save("C:\\Project\\WB_Pars_brands\\new_output.xlsx")

    def all_scrap_func(self, driver, link, ws):
        self.load_page(driver, link)
        self.scroll_page(driver)
        self.settings_fro_execel_file(ws)
        html = driver.page_source
        if html:
            self.scrap_wb(html, ws)

    def scrap_wb(self, html, ws):
        soup = BeautifulSoup(html, 'html.parser')
        cards = soup.find_all("div", class_="product-card j-card-item")

        for card in cards:
            price_tag = card.find("span", class_="price__lower-price")
            price_pattern = re.compile(r'\d[\d\s]*\d')
            price = self.price_scrap(card, price_pattern, price_tag)

            brand_name = self.brand_scrap(card)

            name_tag = self.name_scrap(card)

            count = self.stars_scrap(card)

            href = self.link_scrap(card)

            ws.append([self.number, brand_name, name_tag, price, count, href])
            self.number += 1

    def link_scrap(self, card):
        href_pattern = re.compile(r'href="(.*?)"')
        href = href_pattern.search(card.prettify()).group(1)
        return href

    def stars_scrap(self, card):
        star_tag = card.find("span", class_="product-card__count")
        if star_tag:
            pattern = re.compile(r'\d+')
            match = pattern.search(star_tag.text)
            if match:
                count = match.group(0)
        return count

    def name_scrap(self, card):
        name_tag = card.find("p", class_="product-card__brand-name").find("span",
                                                                          class_="goods-name").text.strip().lstrip(
            '/')
        return name_tag

    def brand_scrap(self, card):
        brand_tag = card.find("p", class_="product-card__brand-name")
        if brand_tag:
            brand_name_tag = brand_tag.find("span", class_="brand-name")
            if brand_name_tag:
                brand_name = brand_name_tag.text.strip()
        return brand_name

    def price_scrap(self, card, price_pattern, price_tag):
        if price_tag is not None:
            if price_tag:
                price_text = price_tag.get_text()
                price_match = price_pattern.search(price_text)
                if price_match:
                    price = int(price_match.group().replace('\xa0', '').replace(' ', ''))
        else:
            price_tag = card.find("ins", class_="price__lower-price")
            if price_tag:
                price_text = price_tag.get_text()
                price_match = price_pattern.search(price_text)
                if price_match:
                    price = int(price_match.group().replace('\xa0', '').replace(' ', ''))
        return price

    def settings_fro_execel_file(self, ws):
        name_width = 65
        count_width = 20
        name_letter = get_column_letter(3)  # 3 - это номер колонки "название"
        width_letter = get_column_letter(5)
        ws.column_dimensions[name_letter].width = name_width
        ws.column_dimensions[width_letter].width = count_width

    def scroll_page(self, driver):
        element = driver.find_element(By.CLASS_NAME, 'product-card-list')
        last_element = None
        while True:
            time.sleep(0.2)
            # Найдите все элементы с классом "j-card-item"
            elements = element.find_elements(By.CLASS_NAME, "j-card-item")
            if last_element and last_element == elements[-1]:
                # Если последний элемент не изменился, значит все элементы загружены
                break
            last_element = elements[-1]
            # Прокрутите страницу на один экран вниз
            driver.execute_script("arguments[0].scrollIntoView();", last_element)

    def load_page(self, driver, link):
        driver.get(link)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        html = driver.page_source
        wait = WebDriverWait(driver, 10)
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "product-card-list")))

    def _init_driver(self):
        person = random.choice(self.users)
        persona = {
            'user-agent': person
        }
        options = webdriver.ChromeOptions()
        options.add_argument(f"user-agent={persona['user-agent']}")
        options.add_argument('--blink-settings=imagesEnabled=false')
        options.add_argument('--headless')
        driver_ = Service(executable_path='web_driver_for_selenium\\chromedriver.exe')
        driver = webdriver.Chrome(options=options, service=driver_)
        return driver


pars = WildBerriesScraper()
pars.run()
